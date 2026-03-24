import sys
import os
# 添加项目根目录到 sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Protection
from openpyxl.formatting.rule import FormulaRule
from models import init_db, get_session, ChannelCustomer, Supplier, Point, SKU, ExternalPartner
from logic.constants import PointType, SupplierCategory, SKUType, ExternalPartnerType

def apply_protection_and_formatting(ws, df, op_col_letter):
    ws.protection.sheet = True # 开启工作表保护，默认全锁定
    ws.protection.formatColumns = False # 允许用户自由调整列宽
    
    # 将默认列宽拉宽一点，以防字被盖住
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 22
        
    # 遍历解锁除了第一列(系统ID)以外的数据单元格
    max_col = len(df.columns)
    for row in ws.iter_rows(min_row=2, max_row=1000, min_col=2, max_col=max_col):
        for cell in row:
            cell.protection = Protection(locked=False)
            
    # 条件格式：选删除时整行变红
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    # 这里我们用绝对引用 $D2="删除"，这样 D 列变化时该行的 A到D都会被标红
    rule = FormulaRule(formula=[f'${op_col_letter}2="删除"'], stopIfTrue=True, fill=red_fill)
    ws.conditional_formatting.add(f"A2:{op_col_letter}1000", rule)

def export_master_data_to_excel():
    # 连接生产数据库
    db_uri = 'sqlite:///data/business_system.db'
    init_db(db_uri)
    session = get_session()
    
    # 准备导出路径
    output_dir = r"d:\WorkSpace\ShanYin\ShanYinERP-v3\docs"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    file_path = os.path.join(output_dir, "master_data_export_full_validation.xlsx")
    
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # ==========================================
            # 1. 渠道客户 
            # ==========================================
            customers = session.query(ChannelCustomer).all()
            cust_data = []
            for c in customers:
                cust_data.append({
                    "[系统ID]": c.id,
                    "客户名称": c.name,
                    "整体信息描述": c.info,
                    "[操作指令]": ""
                })
            df_cust = pd.DataFrame(cust_data) if cust_data else pd.DataFrame(columns=["[系统ID]", "客户名称", "整体信息描述", "[操作指令]"])
            df_cust.to_excel(writer, sheet_name="渠道客户", index=False)
            
            ws_cust = writer.sheets["渠道客户"]
            
            # 操作指令下拉验证
            dv_op_cust = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
            ws_cust.add_data_validation(dv_op_cust)
            dv_op_cust.add("D2:D1000")
            
            apply_protection_and_formatting(ws_cust, df_cust, "D")

            # 隐藏配置页
            if "Settings" not in writer.book.sheetnames:
                ws_settings = writer.book.create_sheet('Settings')
                ws_settings.sheet_state = 'hidden'
                ws_settings['A1'] = "闪饮自身"

            # ==========================================
            # 2. 供应商 
            # ==========================================
            suppliers = session.query(Supplier).all()
            supp_data = []
            for s in suppliers:
                supp_data.append({
                    "[系统ID]": s.id,
                    "供应商名称": s.name,
                    "供应类别": s.category,
                    "地址信息": s.address,
                    "[操作指令]": ""
                })
            df_supp = pd.DataFrame(supp_data) if supp_data else pd.DataFrame(columns=["[系统ID]", "供应商名称", "供应类别", "地址信息", "[操作指令]"])
            df_supp.to_excel(writer, sheet_name="供应商", index=False)
            
            ws_supp = writer.sheets["供应商"]
            
            # 类别下拉验证
            dv_supp_cat = DataValidation(type="list", formula1=f'"{",".join(SupplierCategory.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
            ws_supp.add_data_validation(dv_supp_cat)
            dv_supp_cat.add("C2:C1000")
            
            # 操作指令下拉验证
            dv_op_supp = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
            ws_supp.add_data_validation(dv_op_supp)
            dv_op_supp.add("E2:E1000")
            
            apply_protection_and_formatting(ws_supp, df_supp, "E")
            
            # ==========================================
            # 3. 点位
            # ==========================================
            points = session.query(Point).all()
            point_data = []
            for p in points:
                owner_type = "[自身] 公司"
                owner_name = "闪饮自身"
                if p.customer:
                    owner_type = "[客户]"
                    owner_name = p.customer.name
                elif p.supplier:
                    owner_type = "[供应商]"
                    owner_name = p.supplier.name
                    
                point_data.append({
                    "[系统ID]": p.id,
                    "归属主体类别": owner_type,
                    "归属主体名称": owner_name,
                    "点位名称": p.name,
                    "点位类型": p.type,
                    "详细地址": p.address,
                    "收货地址": p.receiving_address,
                    "[操作指令]": ""
                })
            df_point = pd.DataFrame(point_data) if point_data else pd.DataFrame(columns=["[系统ID]", "归属主体类别", "归属主体名称", "点位名称", "点位类型", "详细地址", "收货地址", "[操作指令]"])
            df_point.to_excel(writer, sheet_name="点位", index=False)
            
            ws_point = writer.sheets["点位"]
            
            # 归属主体类别下拉验证
            dv_owner_type = DataValidation(type="list", formula1='"[自身] 公司,[客户],[供应商]"', allow_blank=True, showErrorMessage=True)
            ws_point.add_data_validation(dv_owner_type)
            dv_owner_type.add("B2:B1000")

            # 归属主体名称下拉验证 (动态联动：引用其他页签的名称列)
            dv_owner_name = DataValidation(type="list", formula1='=IF(B2="[客户]",\'渠道客户\'!$B$2:$B$1000,IF(B2="[供应商]",\'供应商\'!$B$2:$B$1000,\'Settings\'!$A$1:$A$1))', allow_blank=True, showErrorMessage=True)
            ws_point.add_data_validation(dv_owner_name)
            dv_owner_name.add("C2:C1000")
            
            # 点位类型下拉验证
            dv_point_type = DataValidation(type="list", formula1=f'"{",".join(PointType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
            ws_point.add_data_validation(dv_point_type)
            dv_point_type.add("E2:E1000")
            
            # 操作指令下拉验证
            dv_op_point = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
            ws_point.add_data_validation(dv_op_point)
            dv_op_point.add("H2:H1000")
            
            apply_protection_and_formatting(ws_point, df_point, "H")
            
            # ==========================================
            # 4. SKU
            # ==========================================
            skus = session.query(SKU).all()
            sku_data = []
            for k in skus:
                sku_data.append({
                    "[系统ID]": k.id,
                    "所属供应商名称": k.supplier.name if k.supplier else "",
                    "SKU名称": k.name,
                    "一级分类": k.type_level1,
                    "型号": k.model,
                    "[操作指令]": ""
                })
            df_sku = pd.DataFrame(sku_data) if sku_data else pd.DataFrame(columns=["[系统ID]", "所属供应商名称", "SKU名称", "一级分类", "型号", "[操作指令]"])
            df_sku.to_excel(writer, sheet_name="SKU", index=False)
            
            ws_sku = writer.sheets["SKU"]
            
            # 所属供应商名称下拉验证 (引用供应商页签的名称列)
            dv_sku_supplier = DataValidation(type="list", formula1='=\'供应商\'!$B$2:$B$1000', allow_blank=True, showErrorMessage=True)
            ws_sku.add_data_validation(dv_sku_supplier)
            dv_sku_supplier.add("B2:B1000")

            # 一级分类下拉验证
            dv_sku_type = DataValidation(type="list", formula1=f'"{",".join(SKUType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
            ws_sku.add_data_validation(dv_sku_type)
            dv_sku_type.add("D2:D1000")
            
            # 操作指令下拉验证
            dv_op_sku = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
            ws_sku.add_data_validation(dv_op_sku)
            dv_op_sku.add("F2:F1000")
            
            apply_protection_and_formatting(ws_sku, df_sku, "F")
            
            # ==========================================
            # 5. 外部合作方
            # ==========================================
            partners = session.query(ExternalPartner).all()
            partner_data = []
            for p in partners:
                partner_data.append({
                    "[系统ID]": p.id,
                    "机构名称": p.name,
                    "机构类别": p.type,
                    "[操作指令]": ""
                })
            df_partner = pd.DataFrame(partner_data) if partner_data else pd.DataFrame(columns=["[系统ID]", "机构名称", "机构类别", "[操作指令]"])
            df_partner.to_excel(writer, sheet_name="外部合作方", index=False)
            
            ws_partner = writer.sheets["外部合作方"]
            
            # 机构类别下拉验证
            dv_partner_type = DataValidation(type="list", formula1=f'"{",".join(ExternalPartnerType.ALL_TYPES)}"', allow_blank=True, showErrorMessage=True)
            ws_partner.add_data_validation(dv_partner_type)
            dv_partner_type.add("C2:C1000")
            
            # 操作指令下拉验证
            dv_op_partner = DataValidation(type="list", formula1='"新增,更新,删除"', allow_blank=True, showErrorMessage=True)
            ws_partner.add_data_validation(dv_op_partner)
            dv_op_partner.add("D2:D1000")
            
            apply_protection_and_formatting(ws_partner, df_partner, "D")
            
        print(f"✅ 成功导出带有数据验证（下拉菜单）的生产环境基础数据至: {file_path}")
    except Exception as e:
        print(f"❌ 导出失败: {e}")
    finally:
        session.close()

if __name__ == "__main__":
    export_master_data_to_excel()
